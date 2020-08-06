"""
解析数据文件 csv excel
"""
import csv
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import os
import configparser


def do_cvs(csvpath):
    test_data = []
    with open(csvpath, mode='r', encoding='utf8') as f:
        reader = csv.reader(f)
        next(reader)
        for line in reader:
            print(line)
            test_data.append(tuple(line))
    return test_data


def do_excel(filename, sheetname, minrow, maxrow, mincol, maxcol):
    father_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    filename = os.path.join(father_path + "/Data", filename)
    print(filename)
    workbook = load_workbook(filename)
    ws: Worksheet = workbook[sheetname]

    testdata = []
    for row in ws.iter_rows(min_row=minrow, max_row=maxrow, min_col=mincol, max_col=maxcol, values_only=True):
        testdata.append(row)

    return testdata


if __name__ == '__main__':
    data = do_excel(filename="data.xlsx", sheetname="用户登录", minrow=2, maxrow=3, mincol=1, maxcol=3)
    print(data[0][0])
